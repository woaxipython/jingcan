from datetime import datetime

import xlrd
from openpyxl.reader.excel import load_workbook


class StoreProFile():
    def __init__(self, save_path):
        self.save_path = save_path
        self.messages = []
        if "xlsx" in self.save_path:
            self.validate_xlsx()
        elif "xls" in self.save_path:
            self.validate_xls()

    def validate_xls(self):
        workbook = xlrd.open_workbook(self.save_path)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            self.messages.append('表格数据不能为空')
            return False
        elif sheet.ncols != 6:
            self.messages.append('表格标头不正确，请下载模板')
            return False
        elif sheet.cell(0, 0).value != '推广平台' or sheet.cell(0, 1).value != '推广方式' or sheet.cell(0,
                                                                                                        2).value != '推广店铺' \
                or sheet.cell(0, 3).value != '推广产品' or sheet.cell(0, 4).value != '推广花费' or sheet.cell(0,
                                                                                                              5).value != '推广时间':
            self.messages.append('表格列名不正确')
            return False
        else:
            for row in range(1, sheet.nrows):
                if not sheet.cell(row, 0).value or not sheet.cell(row, 1).value or not sheet.cell(row, 2).value or not \
                        sheet.cell(row, 3).value or not sheet.cell(row, 4).value or not sheet.cell(row, 5).value:
                    self.messages.append('表格数据不能为空')
                    break
            for row in range(1, sheet.nrows):
                if not isinstance(sheet.cell(row, 4).value, float):
                    self.messages.append('推广花费应当是数字')
                    break
            for row in range(1, sheet.nrows):
                if not isinstance(sheet.cell(row, 5).value, datetime):
                    self.messages.append('推广时间应当是日期')
                    break
            for row in range(1, sheet.nrows):
                if not isinstance(sheet.cell(row, 2).value, str):
                    self.messages.append('推广店铺应当是字符串')
                    break
            for row in range(1, sheet.nrows):
                if not isinstance(sheet.cell(row, 3).value, str):
                    self.messages.append('推广产品应当是字符串')
                    break
            for row in range(1, sheet.nrows):
                if not isinstance(sheet.cell(row, 0).value, str):
                    self.messages.append('推广平台应当是字符串')
                    break
            for row in range(1, sheet.nrows):
                if not isinstance(sheet.cell(row, 1).value, str):
                    self.messages.append('推广方式应当是字符串')
                    break

    def validate_xlsx(self):
        workbook = load_workbook(filename=self.save_path)
        sheet = workbook.active

        header = ['推广平台', '推广方式', '推广店铺', '推广产品', '推广花费', '推广时间', ]
        head_row = sheet[1]
        for cell in head_row:
            if cell.value != header[head_row.index(cell)]:
                self.messages.append('表头错误')
                break
        for row in sheet.iter_rows(min_row=2):
            if not row[0].value or not row[1].value or not row[2].value or not row[3].value or not row[4].value or not \
                    row[5].value:
                self.messages.append('表格数据错误')
                break
        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[4].value, int):
                self.messages.append('推广花费应当是数字')
                break
        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[5].value, datetime):
                self.messages.append('推广时间应当是日期')
                break

        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[2].value, str):
                self.messages.append('推广店铺应当是字符串')
                break

        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[3].value, str):
                self.messages.append('推广产品应当是字符串')
                break

        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[1].value, str):
                self.messages.append('推广方式应当是字符串')
                break
        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[0].value, str):
                self.messages.append('推广平台应当是字符串')
                break

    def validate(self):
        if self.messages:
            return False
        else:
            return True


class HandOrderFile():
    def __init__(self, save_path):
        self.save_path = save_path
        self.messages = []
        self.validate_xlsx()

    def validate_xlsx(self):
        workbook = load_workbook(filename=self.save_path)
        try:
            sheet = workbook['手工单模板']
        except:
            self.messages.append('表格名称不正确，请下载模板')
            return False

        header = ["订单分类", "发货原因(分销则写分销商名称)", "商品名称", "商品数量", "商品单价", "收件人地址",
                  "下单时间", "快递公司", "快递单号"]

        head_row = sheet[1]
        for cell in head_row:
            if cell.value != header[head_row.index(cell)]:
                self.messages.append('表头错误')
                break
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row - 1):
            if not row[0].value or not row[1].value or not row[2].value or not row[3].value or not row[4].value or not \
                    row[5].value or not row[6].value:
                if row[0].value != 0 and row[1].value != 0 and row[2].value != 0 and row[3].value != 0 and row[
                    4].value != 0 and row[5].value != 0 and row[6].value != 0:
                    self.messages.append('第{}行表格数据不能为空'.format(row[0].row))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[0].value:
                if "手工" not in row[0].value and "分销" not in row[0].value:
                    self.messages.append('第{}行第{}列订单分类错误'.format(row[0].row, row[0].column))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[1].value:
                if not isinstance(row[1].value, str):
                    self.messages.append('第{}行第{}列发货原因应当是字符串'.format(row[0].row, row[0].column))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[2].value:
                if not isinstance(row[2].value, str):
                    self.messages.append('第{}行第{}列商品名称应当是字符串'.format(row[0].row, row[0].column))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[3].value:
                if not isinstance(row[3].value, int):
                    self.messages.append('第{}行第{}列数量应当是整数数字'.format(row[1].row, row[1].column))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[4].value or row[5].value == 0:
                if not isinstance(row[4].value, float) and not isinstance(row[4].value, int):
                    self.messages.append('第{}行第{}列单价应当是数字'.format(row[2].row, row[2].column))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[5].value:
                if not isinstance(row[5].value, str):
                    self.messages.append('第{}行第{}列收件人应当是字符串'.format(row[3].row, row[3].column))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[6].value:
                if not isinstance(row[6].value, datetime):
                    self.messages.append('第{}行第{}列下单时间应该为时间'.format(row[3].row, row[3].column))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[7].value:
                if not isinstance(row[7].value, str):
                    self.messages.append('第{}行第{}列快递公司应该为字符串'.format(row[3].row, row[3].column))
                    return False
        for row in sheet.iter_rows(min_row=2):
            if row[8].value:
                if not isinstance(row[8].value, str):
                    self.messages.append('第{}行第{}列快递单号应该为字符串'.format(row[3].row, row[3].column))
                    return False

    def validate(self):
        if self.messages:
            return False
        else:
            return True


class codeContractFileForm():
    def __init__(self, save_path):
        self.save_path = save_path
        self.messages = []
        self.validate_xlsx()

    def validate_xlsx(self):
        workbook = load_workbook(filename=self.save_path)
        try:
            sheet = workbook.active
        except:
            self.messages.append('表格名称不正确，请下载模板')
            return False
        header = ["销售名称", "商品编码"]
        head_row = sheet[1]
        for cell in head_row:
            if cell.value != header[head_row.index(cell)]:
                self.messages.append('表头错误')
                break
        for row in sheet.iter_rows(min_row=2):
            if not row[0].value or not row[1].value:
                self.messages.append('表格数据不能为空')
        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[0].value, str):
                self.messages.append('第{}行第{}列销售名称应当是字符串'.format(row[0].row, row[0].column))
                return False
        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[1].value, str):
                self.messages.append('第{}行第{}列商品编码应当是字符串'.format(row[1].row, row[1].column))
                return False

    def validate(self):
        if self.messages:
            return False
        else:
            return True


class PromotionFileForm():
    def __init__(self, save_path):
        self.save_path = save_path
        self.messages = []
        self.validate_xlsx()

    def validate_xlsx(self):
        workbook = load_workbook(filename=self.save_path)
        try:
            sheet = workbook.active
        except:
            self.messages.append('表格名称不正确，请下载模板')
            return False
        header = ["推广人", "博主微信", "账号主页链接", "*平台", "*推广产品，多个产品,隔开", "付费形式", "费用",
                  "佣金", "*图文链接", "产出形式", "合作时间", "*账号自营"]
        head_row = sheet[1]
        for cell in head_row:
            if cell.value != header[head_row.index(cell)]:
                self.messages.append('表头错误')
                break
        for row in sheet.iter_rows(min_row=2):
            if not row[3].value:
                self.messages.append('第{}行第{}列平台不能为空'.format(row[3].row, row[3].column))
                return False
        for row in sheet.iter_rows(min_row=2):
            if not row[4].value:
                self.messages.append('第{}行第{}列商品不能为空'.format(row[4].row, row[4].column))
                return False
        for row in sheet.iter_rows(min_row=8):
            if not row[8].value:
                self.messages.append('第{}行第{}列账号图文链接不能为空'.format(row[8].row, row[8].column))
                return False
        for row in sheet.iter_rows(min_row=2):
            if not row[11].value:
                self.messages.append('第{}行第{}列账号自营不能为空'.format(row[11].row, row[11].column))
                return False

        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[3].value, str):
                self.messages.append('第{}行第{}列平台应当是字符串'.format(row[3].row, row[3].column))
                return False
        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[8].value, str) and row[8].value != None:
                self.messages.append('第{}行第{}列图文链接应当是字符串'.format(row[8].row, row[8].column))
                return False
        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[8].value, str):
                self.messages.append('第{}行第{}列图文链接应当是字符串'.format(row[3].row, row[3].column))
                return False
        for row in sheet.iter_rows(min_row=2):
            if not isinstance(row[11].value, str) and row[8].value != None:
                self.messages.append('第{}行第{}列账号自营应当是字符串'.format(row[8].row, row[8].column))
                return False

    def validate(self):
        if self.messages:
            return False
        else:
            return True


if __name__ == '__main__':
    print(HandOrderFile('/Users/zhengyansheng/Desktop/分销商订单模板.xlsx').validate())
