import hashlib
import re
import time
import random
from datetime import datetime

import requests
from flask import jsonify
from openpyxl.reader.excel import load_workbook
from sqlalchemy import func, or_

from APP.SQLAPP.addEdit.dataWrite import generate_Number_string
from APP.SQLAPP.search.promotion import searchAccount, searchNotes
from APP.Spyder.makeRealURL import MakeRealURL
from exts import db
from models.back import RateModel, PlatModel, FeeModel, OutputModel
from models.product import GroupModel
from models.promotion import PromotionModel, BlogerModel, AccountModel, AccountDayDataModel
from models.promotiondata import PVContentModel, PVDataModel
from models.store import OrderModel, ParentOrderModel
from models.user import UserModel


class writePromotionFee(object):
    def __init__(self, form_dict, save_path=""):
        """传入的form_dict前端传入的数据，字典格式，包含了所有的数据"""
        """
        editPromotionId:    ——>  promotion_id 推广id int
        editPromotionOrderId:  ——>  order_id 订单id str
        editPromotionCheck:  ——>  orderstatus 订单状态 str "0" or "1"
        save_path:  ——>  feeImgLink 保存路径 str
        """
        self.form_dict = form_dict
        self.error_message = []
        self.save_path = save_path
        self.search_id = self.form_dict.get("editPromotionId")
        self.order_id = self.form_dict.get("editPromotionOrderId")
        self.order_status = self.form_dict.get("editPromotionCheck")

    def check(self):
        """检查数据是否合法"""
        promotion_model = PromotionModel.query.filter_by(search_id=self.search_id).first()
        if not promotion_model:
            self.error_message.append("没有该推广记录，可尝试刷新页面")
        if self.order_status == "1":
            order_model = OrderModel.query.filter_by(orderID=self.order_id).first()
            if not order_model:
                self.error_message.append("没有该订单")
        if self.error_message:
            return False
        else:
            return True

    def write(self):
        """写入数据"""
        pro_model = PromotionModel.query.filter_by(search_id=self.search_id).first()  # 推广记录
        rate_id = pro_model.rate_id
        order_model = ParentOrderModel.query.filter_by(
            orderID=self.order_id).first() if self.order_status == "1" else None
        rate_model = RateModel.query.filter_by(name="已付款").first()  # 进度模型
        if not rate_model:  # 如果没有已付款的进度模型，创建一个
            rate_model = RateModel(name="已付款")
            pro_model.rate = rate_model
        elif rate_model.id > rate_id:  # 如果已付款的进度模型id大于推广记录的进度模型id，修改推广记录的进度模型
            pro_model.rate = rate_model
        pro_model.feeImgLink = self.save_path  # 保存路径
        pro_model.orderstatus = self.order_status  # 订单状态
        pro_model.order = order_model  # 订单
        db.session.add(pro_model)  # 添加到数据库
        db.session.commit()  # 提交到数据库
        return {"status": "success", "message": "修改成功"}


class writeNewPromotionModel(object):
    def __init__(self, form_dict):
        """传入的form_dict前端传入的数据，字典格式，包含了所有的数据"""
        """
        promotionPr                         ——>商品组，多选，列表，id
        promotionUser                       ——>联络人，id
        promotionRate                       ——>进度模型，id
        promotionFeeModel                   ——>费用模板，id
        promotionWechat                     ——>博主微信，str
        promotionCommission                 ——>佣金比例，float
        PromotionCheck                      ——>订单状态，str
        AccountList                         ——>推广账号列表，列表，字典，id，str
        promotionFee                        ——>费用，float    
        promotionCommission                 ——>佣金，float
        PromotionCheck                      ——>是否下单，str
        """
        self.form_dict = form_dict
        self.error_message = []
        self.pr = self.form_dict.get("promotionPr")
        self.user_id = self.form_dict.get("promotionUser")
        self.rate_id = self.form_dict.get("promotionRate")
        self.fee_model_id = self.form_dict.get("promotionFeeModel")
        self.wechat = self.form_dict.get("promotionWechat")

    def check(self):
        """检查数据是否正确"""
        group_model = GroupModel.query.filter(or_(GroupModel.id.in_(self.pr), GroupModel.name.in_(self.pr))).all()
        if not group_model:
            self.error_message.append("没有该商品组")  # 没有该商品组
        user_model = UserModel.query.filter_by(id=self.user_id).first()
        if not user_model:
            self.error_message.append("没有该联络人")  # 没有该联络人
        rate_model = RateModel.query.filter_by(id=self.rate_id).first()
        if not rate_model:
            self.error_message.append("没有该进度模型")  # 没有该进度模型
        fee_model = FeeModel.query.filter_by(id=self.fee_model_id).first()
        if not fee_model:
            self.error_message.append("没有该费用模板")  # 没有该费用模板
        bloger_model = BlogerModel.query.filter_by(wechat=self.wechat).first()
        if not bloger_model:
            bloger_model = BlogerModel(wechat=self.wechat)  # 没有该博主,创建博主
            db.session.add(bloger_model)
            db.session.commit()
        if self.error_message:
            return False
        else:
            return True

    def write(self, create_time=None):
        group_model = GroupModel.query.filter(
            or_(GroupModel.id.in_(self.pr), GroupModel.name.in_(self.pr))).all()  # 商品组
        bloger_model = BlogerModel.query.filter_by(wechat=self.wechat).first()  # 博主

        max_id = db.session.query(func.max(PromotionModel.id)).scalar() or 0  # 获取最大的id
        search_id = generate_Number_string(8, max_id + 1)  # 生成8位的id
        promotion_model = PromotionModel(search_id=search_id, fee=self.form_dict['promotionFee'],
                                         commission=self.form_dict['promotionCommission'],
                                         orderstatus=self.form_dict['PromotionCheck'],
                                         feeModel_id=self.fee_model_id, rate_id=self.rate_id, user_id=self.user_id,
                                         bloger=bloger_model)  # 创建推广模型
        promotion_model.group = group_model  # 商品组
        if create_time:
            promotion_model.createtime = create_time
        note_search_id = ""  # 生成推广内容的id
        for account_dict in self.form_dict['AccountList']:
            output_model = OutputModel.query.get(account_dict['outputModel'])  # 获取输出模型
            account_model = AccountModel.query.filter_by(account_id=account_dict['account']).first()  # 获取推广账号
            if not account_model:  # 如果没有该推广账号，就返回错误，提示先添加
                return {"status": "failed", "message": "数据库中没有该推广账号，请先添加"}
            account_model.bloger = bloger_model  # 将推广账号的博主设置为该博主
            max_id = db.session.query(func.max(PVContentModel.id)).scalar() or 0  # 获取最大的id
            note_search_id = generate_Number_string(8, max_id + 1)  # 生成8位的id
            new_PV_model = PVContentModel(account=account_model, output=output_model, search_id=note_search_id,
                                          promotion=promotion_model)  # 创建推广内容模型
            db.session.add(new_PV_model)  # 添加到数据库
        db.session.add(promotion_model)  # 添加到数据库
        db.session.commit()  # 提交
        return {"status": "success", "message": "新增成功", 'search_id': search_id, "note_search_id": note_search_id}


makeRealURL = MakeRealURL()


# 写入EXCEL批量导入的推广内容
class WriteExcelPVContent(object):
    def __init__(self, save_path):
        """初始化"""
        """ header = 
        ["推广产品", "*图文链接", "账号自营"]
        """
        self.save_path = save_path
        self.workbook = load_workbook(save_path)
        self.sheet = self.workbook.active
        self.data_list = []
        self.error_message = []

        self.account_dict = {}
        self.account_form = {}
        self.promotion_form = {}
        self.note_dict = {}
        self.readExcel()

    def readExcel(self):  # 读取推广表格
        workbook = load_workbook(self.save_path)  # 打开excel文件
        sheet = workbook.active  # 获取当前活跃的sheet,默认是第一个sheet
        self.total = 0
        for row in sheet.rows:  # 逐行读取
            if row[0].row == 1:  # 跳过第一行
                continue
            row_list = [cell.value for cell in row]  # 读取每行的数据
            self.data_list.append(row_list)  # 读取每行的数据
            self.total += 1

    def write(self):
        """写入文件"""
        i = 0

        for row in self.data_list:
            self.writrNote(row)  # 写入推广内容
            i += 1
            print("------------")
            print("已经写入：{}行，共计无误链接{}行，有误链接{}行；本次图文链接：{}".format(i, len(self.data_list),
                                                                                       len(self.error_message), row[1]))
        print(self.error_message)

    def check(self):
        """检查数据"""
        '["推广产品", "*图文链接", "账号自营"]'
        new_data_list = []
        self.checked = 1  # 因为已经跳过第一行，所以计数从2开始
        for row in self.data_list:
            print("正在检查第{}行，共计预计导入{}行".format(self.checked, self.total))
            self.checked += 1
            if row[0]:
                if GroupModel.query.filter_by(name=row[0]).first() is None:  # 检查商品组是否存在
                    self.error_message.append(f"第{self.checked}行的商品组不存在")
                    continue
            if row[1]:  # 图文链接是否为空，为空则跳过
                pv_url = makeRealURL.makePVContentURL(row[1])  # 重构图文链接为标准格式
                if not pv_url:
                    self.error_message.append(f"第{self.checked}行的图文链接{row[1]}转换错误")
                    continue
                row[1] = pv_url

            new_data_list.append(row)  # 将检查通过的数据添加到新列表中
        self.data_list = new_data_list  # 将检查通过的数据赋值给原来的列表

    def writrNote(self, row):
        """写入推广内容"""
        plat = MakeRealURL().makePlatName(row[1])
        plat_model = PlatModel.query.filter_by(name=plat).first()

        content_id = MakeRealURL().makeContentID(row[1])
        goods_model = GroupModel.query.filter_by(name=row[0]).first()
        pvcontent_model = PVContentModel.query.filter_by(content_id=content_id).first()
        pvcontent_model = PVContentModel(content_id=content_id) if not pvcontent_model else pvcontent_model
        pvcontent_model.content_link = row[1]
        pvcontent_model.plat = plat_model
        pvcontent_model.goods = goods_model
        db.session.add(pvcontent_model)
        db.session.commit()


# 写入SQL查询的各类数据
class WriteSQLData(object):
    def WriteSqlPVcontentData(self, note_link, notes_Info):
        """写入推广内容"""
        content_id = MakeRealURL().makeContentID(note_link)
        if not content_id:
            return {"status": "failed", "message": "请输入正确的笔记链接"}
        pvcontent_model = PVContentModel.query.filter_by(content_id=content_id).first()
        pvcontent_model = PVContentModel(content_id=content_id) if not pvcontent_model else pvcontent_model
        plat = MakeRealURL().makePlatName(note_link)
        plat_model = PlatModel.query.filter_by(name=plat).first()
        for key, value in notes_Info.items(): setattr(pvcontent_model, key, value)
        pvcontent_model.attention = 0 if pvcontent_model.attention != 1 else 1
        pvcontent_model.contenttype = "视频" if notes_Info["video_link"] else "图文"
        pvcontent_model.plat = plat_model
        UNIQUE_ID = MakeRealURL().makeUniqueDayId(note_link)

        db.session.add(pvcontent_model)
        db.session.commit()

        pvcontent_today_model = PVDataModel.query.filter(PVDataModel.search_id == UNIQUE_ID).first()
        pvcontent_today_model = PVDataModel(search_id=UNIQUE_ID) if not pvcontent_today_model else pvcontent_today_model
        pvcontent_today_model.pvcontent = pvcontent_model

        pvcontent_today_model.liked = notes_Info["liked"]
        pvcontent_today_model.collected = notes_Info["collected"]
        pvcontent_today_model.forwarded = notes_Info["forwarded"]
        pvcontent_today_model.commented = notes_Info["commented"]

        account_model = AccountModel.query.filter_by(account_id=notes_Info["account_plat_id"]).first()
        account_model = AccountModel(account_id=notes_Info["account_plat_id"]) if not account_model else account_model
        account_model.attention = 2 if account_model.attention != 1 else 1
        account_model.self = "达人" if account_model.self != "自营" else "自营"
        account_model.profile_link = notes_Info["profile_link"]
        account_model.plat = plat_model

        pvcontent_model.account = account_model

        db.session.add(account_model)

        db.session.add(pvcontent_today_model)
        db.session.commit()

    def writeAccountNotes(self, profile_link, notes_Info):
        account_id = MakeRealURL().makeAccountID(profile_link)
        account_model = AccountModel.query.filter_by(account_id=account_id).first()

        plat = MakeRealURL().makePlatName(profile_link)
        plat_model = PlatModel.query.filter_by(name=plat).first()

        note_link = notes_Info["content_link"]
        content_id = MakeRealURL().makeContentID(note_link)
        pvcontent_model = PVContentModel.query.filter(PVContentModel.content_id == content_id).first()
        pvcontent_model = PVContentModel(content_id=content_id) if not pvcontent_model else pvcontent_model

        for key, value in notes_Info.items(): setattr(pvcontent_model, key, value)
        pvcontent_model.account = account_model
        pvcontent_model.plat = plat_model
        pvcontent_model.attention = 2

        DaydataID = MakeRealURL().makeUniqueDayId(note_link)
        pvcontent_today_model = PVDataModel.query.filter(PVDataModel.search_id == DaydataID).first()
        pvcontent_today_model = PVDataModel() if not pvcontent_today_model else pvcontent_today_model
        pvcontent_today_model.search_id = DaydataID
        pvcontent_today_model.pvcontent = pvcontent_model
        pvcontent_today_model.liked = notes_Info["liked"]
        pvcontent_today_model.collected = notes_Info["collected"]
        pvcontent_today_model.forwarded = notes_Info["forwarded"]
        pvcontent_today_model.commented = notes_Info["commented"]

        db.session.add(pvcontent_model)
        db.session.add(pvcontent_today_model)
        db.session.commit()

    def changeSQLNoteStatus(self, note_link, status):
        pvcontent_model = PVContentModel.query.filter_by(content_link=note_link).first()
        if not pvcontent_model:
            return {"status": "failed", "message": "没有该推广内容"}
        pvcontent_model.status = status
        db.session.add(pvcontent_model)
        db.session.commit()

    def changeSQLAccountStatus(self, note_link, status):
        account_model = AccountModel.query.filter_by(profile_link=note_link).first()
        if not account_model:
            return {"status": "failed", "message": "没有该账号"}
        account_model.status = status
        db.session.add(account_model)
        db.session.commit()

    def WriteSqlAccount(self, profile_link, account_Info, plat, selfs=""):
        """写入推广账号"""
        account_id = MakeRealURL().makeAccountID(profile_link)
        print(account_id)
        account_model = AccountModel.query.filter(AccountModel.account_id == account_id).first()
        account_model = AccountModel(account_id=account_id) if not account_model else account_model

        UNIQUE_ID = MakeRealURL().makeUniqueDayId(profile_link)
        plat_model = PlatModel.query.filter_by(name=plat).first()
        for key, value in account_Info.items(): setattr(account_model, key, value)
        account_model.attention = 0 if account_model.attention != 1 else 1
        selfs = selfs if selfs else ""
        account_model.plat = plat_model
        account_model.self = selfs

        account_Day = AccountDayDataModel.query.filter(AccountDayDataModel.search_id == UNIQUE_ID).first()
        account_Day = AccountDayDataModel(search_id=UNIQUE_ID) if not account_Day else account_Day
        account_Day.account = account_model
        account_Day.nickname = account_Info["nickname"]
        account_Day.fans = account_Info["fans"]
        account_Day.notes = account_Info["notes"]
        account_Day.liked = account_Info["liked"]
        account_Day.collected = account_Info["collected"]
        account_Day.follow = account_Info["follow"]

        db.session.add(account_model)
        db.session.add(account_Day)
        db.session.commit()
