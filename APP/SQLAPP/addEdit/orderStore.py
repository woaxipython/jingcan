import hashlib
import os
from datetime import datetime, timedelta
import xlwt
from flask import jsonify, current_app
from sqlalchemy import func, and_, or_

from APP.SQLAPP.addEdit.dataWrite import generate_Number_string
from APP.Spyder.KdzsSpyder import KuaiDiZhuShouSpyder
from APP.Spyder.getAddress import getAddress
from APP.createCodeId import createOrderCode
from exts import db
from models.back import PlatModel, AdMethodModel, ProvinceModel, CityModel, CountyModel
from models.product import SaleModel, GroupModel, AtomSalesModel
from models.store import StoreModel, ParentOrderModel, OrderModel, RefundModel, DistributionModel, HandOrderModel, \
    HandOrderCategory, HandParentOrderModel, AdFeeModel, CodeStractModel
from models.user import UserModel
import pandas as pd
from openpyxl.reader.excel import load_workbook

kdzs = KuaiDiZhuShouSpyder()


# 写入订单数据
class writeOrderData(object):
    """写入在快递助手爬回来的订单数据"""

    def __init__(self, order_info):
        self.order_info = order_info
        self.store_id = self.order_info['sellerId']
        self.plat_store_name = self.order_info['sellerNick']
        self.store_name = self.order_info['sellerNick']
        self.orderID = self.order_info['orderID']
        self.province = self.order_info['receiverProvince']
        self.city = self.order_info['receiverCity']
        self.totalPayment = self.order_info['totalPayment']
        self.totalReceivedPayment = self.order_info['totalReceivedPayment']
        self.updateTime = self.order_info['updateTime']
        self.payTime = self.order_info['payTime']
        self.orders = self.order_info['orders']
        self.checkParentOrder()
        self.writrOrder()

    def checkParentOrder(self):
        """查询父订单是否存在"""
        self.parent_order_model = ParentOrderModel.query.filter_by(orderID=self.orderID).first()  # 查询父订单是否存在
        if not self.parent_order_model:  # 如果不存在就创建
            self.parent_order_model = ParentOrderModel(orderID=self.orderID, province=self.province,
                                                       city=self.city,
                                                       totalPayment=self.totalPayment,
                                                       totalReceivedPayment=self.totalReceivedPayment,
                                                       updateTime=self.updateTime)  # 创建父订单
        else:  # 如果存在就更新
            self.parent_order_model.province = self.province
            self.parent_order_model.city = self.city
            self.parent_order_model.totalPayment = self.totalPayment
            self.parent_order_model.totalReceivedPayment = self.totalReceivedPayment
            self.parent_order_model.updateTime = self.updateTime
        self.parent_order_model.payTime = self.payTime

    def writrOrder(self):
        """写入订单"""
        for order in self.orders:  # 遍历订单
            self.order_model = OrderModel.query.filter_by(orderID=order['orderID']).first()  # 查询订单是否存在
            if not self.order_model:  # 如果不存在就创建
                self.order_model = OrderModel(orderID=order['orderID'], code=order['SkuName'],
                                              quantity=order['quantity'],
                                              payment=order['payment'], updateTime=self.updateTime,
                                              express=order['express'], expressOrder=order['expressOrder'])
            self.order_model.refund = order['refund']  # 更新退款
            self.order_model.status = order['status']  # 更新订单
            self.order_model.code = order['SkuName']  # 更新订单
            sale_pr_model = SaleModel.query.filter_by(sale_name=order["SkuName"]).first()  # 查询商品是否存在
            if not sale_pr_model:  # 如果不存在，检查商品对照表是否存在
                constract_model = CodeStractModel.query.filter_by(name=order['SkuName']).first()  # 检查是否已存在于商品对照表
                if not constract_model:  # 如果不存在就创建，同时创建商品以及商品对照表
                    constract_model = CodeStractModel(name=order['SkuName'], store=self.store_model,
                                                      store_title=order['title'], )
                    db.session.add(constract_model)
                else:
                    sale_pr_model = constract_model.sale  # 如果存在就直接将商品对照表中的商品赋值给商品，并且扣减原料库存
            if sale_pr_model:  # 再次判断商品是否存在
                sale_pr_model.store.append(self.store_model)  # 更新商品存在的店铺
                self.order_model.sale = sale_pr_model
                self.order_model.cost = sale_pr_model.cost
                for atom in sale_pr_model.atoms:  # 如果存在，则更新原材料库存
                    # 查询原材料数量
                    quantity_model = AtomSalesModel.query.filter_by(atomid=atom.id, saleid=sale_pr_model.id).first()
                    if quantity_model:
                        # 更新原材料库存
                        atom.quantity = int(atom.quantity) - int(order['quantity']) * int(quantity_model.quantity)
                        db.session.add(atom)

            self.parent_order_model.express = order['express']  # 更新父订单的快递
            self.parent_order_model.expressOrder = order['expressOrder']  # 更新父订单的快递单号
            self.parent_order_model.status = order['status']  # 更新父订单的状态

            self.order_model.parent_order = self.parent_order_model
            db.session.add(self.order_model)
            db.session.commit()


def writeRefund(refund_result):
    """写入快递助手爬回的退款信息"""
    refund_model = RefundModel.query.filter_by(refund_id=refund_result['refundId']).first()  # 查询退款是否存在
    parent_model = ParentOrderModel.query.filter_by(orderID=refund_result['tid']).first()  # 查询父订单是否存在
    parent_model = ParentOrderModel(orderID=refund_result['tid']) if not parent_model else parent_model  # 如果不存在就创建

    store_model = StoreModel.query.filter_by(store_id=refund_result['sellerId']).first()  # 查询店铺是否存在
    store_model = StoreModel(store_id=refund_result['sellerId']) if not store_model else store_model  # 如果不存在就创建

    # 退款订单匹配商品
    sale_pr_model = SaleModel.query.filter_by(code=refund_result["code"]).first()  # 查询商品是否存在
    if not sale_pr_model:  # 如果不存在就创建
        constract_model = CodeStractModel.query.filter_by(name=refund_result['SkuName']).first()  # 检查是否已存在于商品对照表
        if not constract_model:  # 如果不存在就创建，同时创建商品以及商品对照表
            constract_model = CodeStractModel(name=refund_result['SkuName'], store=store_model)  # 创建商品对照表
            db.session.add(constract_model)
        else:
            sale_pr_model = constract_model.sale  # 如果存在就直接将商品对照表中的商品赋值给商品
    if sale_pr_model:  # 再次判断商品是否存在
        sale_pr_model.store.append(store_model)
        refund_model.sale_id = sale_pr_model.id  # 更新退款商品

    if not refund_model:  # 如果不存在就创建退款订单
        refund_model = RefundModel(refund_id=refund_result["refundId"])

    refund_model.status = refund_result["StatusDesc"]  # 更新退款状态
    refund_model.reason = refund_result["reason"]  # 更新退款原因
    refund_model.amount = refund_result["Amount"]  # 更新退款金额
    refund_model.modifiedTime = refund_result["ModifiedTime"]  # 更新退款修改时间
    refund_model.updateTime = refund_result["CreatedTime"]  # 更新退款创建时间

    refund_model.parent_order = parent_model  # 更新退款父订单
    refund_model.store = store_model  # 更新退款店铺
    db.session.add(refund_model)
    db.session.commit()


def writeStore(store_list):
    """写入快递助手爬回的店铺信息"""
    new_store = []
    for store in store_list:
        store_id = store['sellerId']
        store_model = StoreModel.query.filter_by(store_id=store_id).first()
        print(store['sellerAbbreviation'])
        if not store_model:  # 如果不存在就创建
            store_model = StoreModel(store_id=store_id, )  # 创建店铺
            new_store.append(store['sellerAbbreviation'])  # 记录新店铺

        store_model.name = store['sellerAbbreviation']  # 更新店铺名称
        store_model.plat_store_name = store['sellerNick']  # 更新店铺昵称
        store_model.bindTime = store['bindTime']  # 更新店铺绑定时间
        store_model.status = True if store['status'] == 1 else False  # 更新店铺状态

        plat_name = PlatNameZH(store['platform'].upper())  # 更新店铺平台
        plat_model = PlatModel.query.filter_by(name=plat_name).first()  # 查询平台是否存在
        if plat_model:  # 如果存在，就更新平台
            plat_model.EH_name = store['platform']
            plat_model.is_Store = True
        else:  # 如果不存在，就创建平台
            plat_model = PlatModel(name=plat_name, EH_name=store['platform'], is_Store=True)
        store_model.plat = plat_model  # 更新店铺平台
        db.session.add(store_model)  # 添加店铺
        db.session.commit()  # 提交
    return new_store  # 返回新店铺列表


def PlatNameZH(EHname):
    name_dict = {"PDD": "拼多多", "TB": "淘宝", "FXG": "抖音", "JD": "京东", "KSXD": "快手", "OTHER": "其它",
                 "YHD": "一号店", "B2B": "B2B", "B2C": "B2C"
        , "B2B2C": "B2B2C", "XHS": "小红书"}
    return name_dict.get(EHname)


def writeAdMethodModel(form_dict):
    """写入推广方式表"""
    plat_model = PlatModel.query.get(form_dict.get("platName"))  # 查询平台是否存在
    if plat_model:  # 如果存在就创建
        ad_method_model = AdMethodModel.query.filter_by(name=form_dict.get("proName"),
                                                        plat_id=plat_model.id).first()  # 查询推广方式是否存在
        if ad_method_model:  # 如果存在就返回
            return jsonify({"status": "failed", "message": "推广方式已存在"})  # 如果不存在就创建
        else:
            ad_method_model = AdMethodModel(name=form_dict.get("proName"), plat_id=plat_model.id,
                                            desc=form_dict.get("proFeeModel"))  # 创建推广方式
            db.session.add(ad_method_model)  # 添加推广方式
            db.session.commit()  # 提交
            return jsonify({"status": "success", "message": "新增成功"})
    else:
        return jsonify({"status": "failed", "message": "平台不存在"})


def writeNewDisModel(form_dict):
    """写入新分销公司"""
    user_model = UserModel.query.get(form_dict.get("newDisUser"))  # 查询联络人是否存在
    if not user_model:
        return jsonify({"status": "failed", "message": "没有该联络人"})  # 如果不存在就返回
    else:
        dis_model = DistributionModel.query.filter_by(campany_name=form_dict.get("newDisCampany")).first()  # 查询分销公司是否存在
        if dis_model:  # 如果存在就返回
            return jsonify({"status": "failed", "message": "分销公司已存在"})  # 返回已经存在
        else:
            dis_model = DistributionModel(campany_name=form_dict.get("newDisCampany"), name=form_dict.get("newDisName"),
                                          wechat=form_dict.get("newDisWechat"),
                                          channel=form_dict.get("newDisSaleChannel"),
                                          phone=form_dict.get("newDisPhone"), city=form_dict.get("newDisCity"),
                                          province=form_dict.get("newDisProvince"), link=form_dict.get("newDisLink"),
                                          address=form_dict.get("newDisAddress"),
                                          remark=form_dict.get("newDisRemark"))  # 创建分销公司
            dis_model.user = user_model
            db.session.add(dis_model)
            db.session.commit()
            return jsonify({"status": "success", "message": "新增成功"})


class WriteHandOrder(object):
    """写入手工订单"""

    def __init__(self, form_dict):
        self.form_dict = form_dict
        self.error_message = []

    def check(self):
        """检查数据"""
        self.hands_type = HandOrderCategory.query.get(self.form_dict['handOrderModel'])  # 查询手工订单类型
        if not self.hands_type:
            self.error_message.append("没有该手工订单类型")
            return False
        self.user_model = UserModel.query.get(self.form_dict['user'])  # 查询用户
        if not self.user_model:
            self.error_message.append("没有该用户")
            return False
        self.order_list, self.shipInfo, self.payment = kdzs.dealCreateOrder(self.form_dict)  # 处理订单
        for order in self.order_list:
            sale_model = SaleModel.query.filter_by(code=order['outerSkuId']).first()  # 查询商品是否存在
            if not sale_model:
                self.error_message.append("没有该商品")
                return False
        return True

    def uploadKdzsOrder(self):
        """写入手工订单"""
        max_id = db.session.query(func.max(HandParentOrderModel.id)).scalar() or 0  # 查询最大id
        self.search_id = createOrderCode(max_id + 1)  # 生成订单号
        create_handle_order = kdzs.createTrade(form_dict=self.form_dict, order_list=self.order_list,
                                               shipInfo=self.shipInfo, search_id=self.search_id)  # 将订单回传给快递助手

        if create_handle_order['status']:
            return True
        else:
            self.error_message.append(create_handle_order['message'])
            return False  # 如果成功

    def weiteOwnData(self):
        """写入自己的数据"""
        handorder_model = HandParentOrderModel(search_id=self.search_id, province=self.form_dict['handOrderProvince'],
                                               city=self.form_dict['handOrderCity'],
                                               address=self.form_dict['handOrderAddress'],
                                               phone=self.form_dict['handOrderPhone'],
                                               name=self.form_dict['handOrderName'],
                                               payment=self.payment,
                                               remark=self.form_dict['handOrderRemark']
                                               )  # 创建手工订单
        handorder_model.user = self.user_model  # 更新联络人
        handorder_model.category = self.hands_type  # 更新订单类型
        for order in self.order_list:  # 遍历订单列表
            order_model = HandOrderModel(code=order['outerSkuId'], quantity=order['num'], payment=order['payment'])
            sale_model = SaleModel.query.filter_by(code=order['outerSkuId']).first()  # 查询商品
            order_model.sale = sale_model
            order_model.cost = sale_model.cost
            order_model.hand_parent_order = handorder_model
            db.session.add(order_model)
        db.session.add(handorder_model)
        db.session.commit()
        return {"status": "success", "message": "创建订单成功"}


class WriteExcelOrder(object):
    """写入EXCEL手工订单"""

    def __init__(self, save_path):
        self.save_path = save_path  # 保存路径

    def dealHandOrder(self):
        """处理手工订单,生成可以直接写入到数据库的JSON样式"""
        workbook = load_workbook(self.save_path)  # 加载上传的EXCEL
        sheet = workbook['手工单模板']  # 选择工作表
        for index, row in enumerate(sheet.rows):  # 遍历行
            if row[0].value and index != 0:
                sellerId = self.makeHandOrderStore_id(row[0].value)
                OrderId = self.makeHandOrderId(str(row[1].value) + str(row[2].value) + str(row[3].value) + str(
                    row[5].value) + str(row[6].value) + str(row[8].value))

                city = self.makeHandOrderCity(row[5].value)
                province = self.makeHandOrderProvince(row[5].value)

                parent_order = {
                    "sellerId": sellerId,
                    "sellerNick": row[0].value,
                    "platform": "线下订单",
                    "orderID": OrderId,
                    "receiverProvince": province if province else "",
                    "receiverCity": city if city else "",
                    "totalPayment": row[4].value,
                    "totalReceivedPayment": row[4].value,
                    "updateTime": row[6].value,
                    "payTime": row[6].value,
                    'orders': []
                }
                order_dict = {
                    "code": row[2].value,
                    "SkuName": row[2].value,
                    "quantity": row[3].value,
                    "orderID": OrderId,
                    "title": row[2].value,
                    "payment": row[4].value,
                    "refund": "未退款",
                    "status": "交易成功",
                    "express": row[7].value,
                    "expressOrder": row[8].value, }
                parent_order["orders"].append(order_dict)
                yield parent_order

    def makeHandOrderProvince(self, info):
        province_list = [row.name for row in ProvinceModel.query.all()]
        for province in province_list:
            if province.replace("城区", "市") in info:
                return province

    def makeHandOrderCity(self, info):
        city_list = [row.name for row in CityModel.query.all()]
        for city in city_list:
            if city in info:
                return city

    def makeHandOrderStore_id(self, info):
        if "手工" in info:
            return "1251533"
        elif "分销" in info:
            return "168188696251881"
        else:
            return ""

    def makeHandOrderId(self, info):
        encoded_text = hashlib.sha1(info.encode()).hexdigest()
        return encoded_text[:20]

    def makeKdzsExcel(self):
        """生成快递助手手工订单EXCEL"""
        workbook = load_workbook(self.save_path)  # 加载上传的EXCEL
        sheet = workbook['分销商订单模板']  # 选择工作表
        max_id = db.session.query(func.max(HandParentOrderModel.id)).scalar() or 0  # 查询最大id
        search_id = createOrderCode(max_id + 1)  # 生成订单号
        self.order_lists = []  # 订单列表

        columns = ["订单编号", "*收件人", "固话", "*手机", "*地址", "发货信息", "*货品规格编码", "宝贝数量", "总重量",
                   "备注", "代收货款(是/否)", "保价服务(是/否)", "实付金额"]  # 列名
        self.order_lists.append(columns)  # 添加列名
        for row in sheet.rows:  # 遍历行
            if row[0].row == 1:  # 如果是第一行
                continue  # 跳过
            order_list = []  # 订单列表
            name = row[0].value  # 产品名称
            sale_model = SaleModel.query.filter_by(name=name).first()  # 查询产品
            if not sale_model:  # 如果没有该产品
                return {"status": "failed", "message": "没有第{}行产品，请重新填写".format(row[0].column)}  # 返回错误信息
            order_list.append(search_id)  # 订单号
            order_list.append(row[3].value)  # 收件人
            order_list.append("")  # 固话
            order_list.append(row[4].value)  # 手机
            order_list.append(row[5].value)  # 地址
            order_list.append(name + ':' + str(row[1].value) + "只")  # 发货信息
            order_list.append(sale_model.code)  # 货品规格编码
            order_list.append(row[1].value)  # 宝贝数量
            order_list.append("")  # 总重量
            order_list.append("")  # 备注
            order_list.append("")  # 代收货款
            order_list.append("")  # 保价服务
            order_list.append(row[2].value)  # 实付金额
            self.order_lists.append(order_list)  # 添加到订单列表
            search_id = "O" + str(int(search_id[1:]) + 1)  # 生成下一个订单号
        workbook.close()  # 关闭EXCEL
        wb = xlwt.Workbook()  # 创建EXCEL
        ws = wb.add_sheet('sheet1')  # 创建工作表
        i = 0
        for row in self.order_lists:  # 遍历订单列表
            j = 0  # 列
            for col in row:  # 遍历列
                ws.write(i, j, col)  # 写入数据
                j += 1
            i += 1
        wb.save(self.upload_path)  # 保存EXCEL
        return {"status": "success", "message": "生成成功"}

    def uploadExcelFile(self):
        """上传EXCEL到快递助手"""
        result = kdzs.uploadHandOrder(upload_path=self.upload_path)  # 上传EXCEL到快递助手
        print(result)
        if result['status'] == "failed":  # 如果上传失败
            return {"status": "failed", "message": result['message']}  # 返回错误信息
        else:
            return {"status": "success", "message": "上传成功"}  # 返回成功信息

    def writeHandOrder(self):
        """写入手工订单到数据库"""
        for order_list in self.order_lists:
            # 查询用户
            user_model = UserModel.query.get(self.form_dict['disOrderUser'])
            # 查询订单分类
            order_category = HandOrderCategory.query.filter_by(name="分销商订单").first()
            if not order_category:
                order_category = HandOrderCategory(name="分销商订单")
            handorder_model = HandParentOrderModel(search_id=order_list[0], address=order_list[4],
                                                   payment=order_list[12],
                                                   name=order_list[1], phone=order_list[3],
                                                   )
            handorder_model.user = user_model
            handorder_model.category = order_category
            order_model = HandOrderModel(code=order_list[6], quantity=order_list[7], payment=order_list[12])
            sale_model = SaleModel.query.filter_by(code=order_list[6]).first()
            order_model.sale = sale_model
            order_model.hand_parent_order = handorder_model
            db.session.add(order_model)
            db.session.add(handorder_model)
        db.session.commit()
        return {"status": "success", "message": "创建订单成功"}


def writeStorePromotionFee(form_dict):
    """写入店铺促销费用"""
    error_message = []
    for Ad in form_dict['AdList']:
        write_result = writeStoreFee(Ad)
        if write_result['status'] == "failed":
            error_message.append(write_result['message'])
            continue
    if error_message:
        return jsonify({"status": "failed", "message": error_message})
    else:
        return jsonify({"status": "success", "message": "新增广告费用成功"})


class WriteStorePromotionFile(object):
    def __init__(self, save_path):
        self.save_path = save_path
        self.error_message = []
        self.workbook = load_workbook(save_path)
        self.sheet = self.workbook.active

    def write(self):
        for row in self.sheet.rows:
            if row[0].row == 1:
                continue
            plat_model = PlatModel.query.filter_by(name=row[0].value).first()
            if not plat_model:
                self.error_message.append("第" + str(row[0]) + "列的平台不存在")
                continue
            plat_id = plat_model.id
            method_model = AdMethodModel.query.filter_by(plat_id=plat_id, name=row[1].value).first()
            if not method_model:
                self.error_message.append("第" + str(row[1].column) + "列的广告方式不存在")
                continue
            method_id = method_model.id
            store_model = StoreModel.query.filter_by(name=row[2].value).first()
            if not store_model:
                self.error_message.append("第" + str(row[2].column) + "列的店铺不存在")
                continue
            store_id = store_model.id
            group_model = GroupModel.query.filter_by(name=row[3].value).first()
            if not group_model:
                self.error_message.append("第" + str(row[3].column) + "列的商品组不存在")
                continue
            group_id = group_model.id
            Ad = {'plat': plat_id, 'method': method_id, 'store': store_id, 'product': group_id, 'fee': row[4].value,
                  'date': row[5].value, }
            writeStoreFee(Ad)
        return self.error_message


def writeStorePromotionFile(save_path):
    """批量写入店铺促销费用"""
    workbook = load_workbook(save_path)
    sheet = workbook.active
    error_message = []
    for row in sheet.rows:
        if row[0].row == 1:
            continue
        print(row[0].value)
        plat_model = PlatModel.query.filter_by(name=row[0].value).first()
        if not plat_model:
            error_message.append("第" + str(row[0]) + "列的平台不存在")
            continue
        plat_id = plat_model.id
        method_model = AdMethodModel.query.filter_by(plat_id=plat_id, name=row[1].value).first()
        if not method_model:
            error_message.append("第" + str(row[1].column) + "列的广告方式不存在")
            continue
        method_id = method_model.id
        store_model = StoreModel.query.filter_by(name=row[2].value).first()
        if not store_model:
            error_message.append("第" + str(row[2].column) + "列的店铺不存在")
            continue
        store_id = store_model.id
        group_model = GroupModel.query.filter_by(name=row[3].value).first()
        if not group_model:
            error_message.append("第" + str(row[3].column) + "列的商品组不存在")
            continue
        group_id = group_model.id
        Ad = {'plat': plat_id, 'method': method_id, 'store': store_id, 'product': group_id, 'fee': row[4].value,
              'date': row[5].value, }
        write_result = writeStoreFee(Ad)
        if write_result['status'] == "failed":
            error_message.append(write_result['message'])
            continue
    if error_message:
        return jsonify({"status": "failed", "message": error_message})
    else:
        return jsonify({"status": "success", "message": "新增成功"})


def writeStoreFee(Ad):
    plat = Ad['plat']
    method = Ad['method']
    store = Ad['store']
    product = Ad['product']
    ad_model = AdMethodModel.query.filter_by(plat_id=plat, id=method).first()
    if not ad_model:
        return {"status": "failed", "message": "广告方式不存在"}
    store_model = StoreModel.query.filter_by(id=store).first()
    if not store_model:
        return {"status": "failed", "message": "店铺不存在"}
    product_model = GroupModel.query.filter_by(id=product).first()
    if not product_model:
        return {"status": "failed", "message": "商品组不存在"}
    maxid = db.session.query(func.max(AdFeeModel.id)).scalar() or 0 + 1
    search_id = datetime.now().strftime('%Y-%m-%d') + "-" + generate_Number_string(6, maxid)
    fee_model = AdFeeModel.query.filter_by(search_id=search_id).first()
    if not fee_model:
        fee_model = AdFeeModel(search_id=search_id)
    fee_model.admethod = ad_model
    fee_model.store = store_model
    fee_model.group = product_model
    fee_model.upload_time = Ad['date']
    fee_model.fee = Ad['fee']
    db.session.add(fee_model)
    db.session.commit()
    return {"status": "success", "message": "新增成功"}
